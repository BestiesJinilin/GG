import calendar as cal_module
import datetime
import json
import openpyxl


from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import check_password as django_check_password
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import auth, User
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.conf import settings
from django.utils import timezone
from django.utils.timezone import localtime
from django.views.decorators.http import require_POST
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment

# FIX: removed `from dateutil.relativedelta import relativedelta`
# python-dateutil is not in Pipfile and causes ImportError on startup.
# Replaced with the pure-Python helper _add_months() below.

from .forms import (
    BeneficiaryFormSet, BookingForm, ClientForm,
    EmployeeCreateForm, EmployeeUpdateForm, PlanForm,
)
from .models import ActivityLog, Booking, ClientPersonalInfo, ClientStatus, Payment, UserLog




# ─────────────────────────────────────────────── helpers ──────────────────────

def _add_months(dt: datetime.date, months: int) -> datetime.date:
    """
    Pure-Python equivalent of relativedelta(months=n).
    Returns the first day of the month that is `months` after `dt`.
    Example: _add_months(date(2025, 1, 15), 2) → date(2025, 3, 1)
    """
    total_months = dt.month - 1 + months
    year  = dt.year + total_months // 12
    month = total_months % 12 + 1
    return dt.replace(year=year, month=month, day=1)


def _check_pin(user, entered_pin: str) -> bool:
    """Return True when entered_pin matches the correct PIN for user."""
    entered_pin = str(entered_pin).strip()
    if user.is_superuser or user.username == "admin":
        from django.conf import settings
        return entered_pin == str(getattr(settings, "ADMIN_PIN", "0000"))
    log = UserLog.objects.filter(user=user).first()
    if not log:
        return False
    return django_check_password(entered_pin, log.pin)


def _admin_required(view_func):
    from functools import wraps

    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("login")
        if not (request.user.is_superuser or request.user.username == "admin"):
            messages.error(request, "Access denied. Admin only.")
            return redirect("homepage")
        return view_func(request, *args, **kwargs)

    return wrapper

def _require_role(*allowed_roles):
    """
    View decorator: grants access to admin + any listed role.
    """
    from functools import wraps

    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect("login")
            if request.user.is_superuser or request.user.username == "admin":
                return view_func(request, *args, **kwargs)
            log = UserLog.objects.filter(user=request.user).first()
            if log and log.role in allowed_roles:
                return view_func(request, *args, **kwargs)
            messages.error(request, "You don't have permission to access this page.")
            return redirect("homepage")
        return wrapper
    return decorator


def _log_activity(user, action: str, detail: str = ""):
    """
    1. Updates the employee's UserLog.activities (latest action only).
    2. Creates a new ActivityLog entry (full permanent history).
    """
    staff_name = "Admin"
    role       = "Admin"
 
    if user.is_authenticated and not (user.is_superuser or user.username == "admin"):
        log = UserLog.objects.filter(user=user).first()
        if log:
            log.activities = action
            log.save(update_fields=["activities"])
            staff_name = log.full_name or user.username
            role       = log.role or "—"
 
    ActivityLog.objects.create(
        user=user if user.is_authenticated else None,
        staff_name=staff_name,
        role=role,
        action=action,
        detail=detail,
    )


def _check_write_role(request, *allowed_roles) -> bool:
    """
    Returns True if the current user may perform a write operation.
    Admin always returns True. Staff must match one of allowed_roles.
    """
    if request.user.is_superuser or request.user.username == "admin":
        return True
    log = UserLog.objects.filter(user=request.user).first()
    return bool(log and log.role in allowed_roles)

def _get_active_status(client):
    """
    Returns the active ClientStatus for a client.
    Priority: active non-cancelled plan → most recent by pk.
    """
    return (
        ClientStatus.objects
        .filter(client=client, status=True, is_cancelled=False)
        .exclude(plan="No Plan")
        .order_by("-pk")
        .first()
    ) or ClientStatus.objects.filter(client=client).order_by("-pk").first()


def _generate_payment_rows(client_status):
    """
    Create one Payment row per month for a ClientStatus.
    FIX: uses _add_months() instead of dateutil.relativedelta.
    Skips silently if rows already exist (idempotent guard).
    Does nothing for the 'No Plan' placeholder (duration == 0).
    """
    if client_status.duration == 0:
        return
    if client_status.payments.exists():
        return
    first_month = client_status.start_date.replace(day=1)
    Payment.objects.bulk_create([
        Payment(
            client_status=client_status,
            month=_add_months(first_month, i),
            amount=client_status.monthly_payment,
            is_paid=False,
        )
        for i in range(client_status.duration)
    ])


# ─────────────────────────────────────────────── auth ─────────────────────────

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "")
        password = request.POST.get("password", "")
        user = auth.authenticate(username=username, password=password)

        if user is not None:
            auth.login(request, user)
            log = UserLog.objects.filter(user=user).first()
            if log:
                log.time_in    = timezone.now()
                log.time_out   = None
                log.activities = "Login"
                log.save()
            # ADD THIS — create activity log entry for login
            staff_name = "Admin"
            role       = "Admin"
            if not (user.is_superuser or user.username == "admin"):
                ulog = UserLog.objects.filter(user=user).first()
                if ulog:
                    staff_name = ulog.full_name or user.username
                    role       = ulog.role or "—"
            ActivityLog.objects.create(
                user=user,
                staff_name=staff_name,
                role=role,
                action="Login",
                detail="",
            )
            return redirect("homepage")

        messages.error(request, "Invalid username or password.")
        return redirect("login")

    return render(request, "app/login.html")


@login_required(login_url="login")
@require_POST
def logout(request):
    if not (request.user.is_superuser or request.user.username == "admin"):
        log = UserLog.objects.filter(user=request.user).first()
        if log:
            log.time_out   = timezone.now()
            log.activities = "Logout"
            log.save()
        # ADD THIS — create activity log entry for logout
    staff_name = "Admin"
    role       = "Admin"
    if not (request.user.is_superuser or request.user.username == "admin"):
        ulog = UserLog.objects.filter(user=request.user).first()
        if ulog:
            staff_name = ulog.full_name or request.user.username
            role       = ulog.role or "—"
    ActivityLog.objects.create(
        user=request.user,
        staff_name=staff_name,
        role=role,
        action="Logout",
        detail="",
    )
    auth.logout(request)
    return redirect("login")


# ─────────────────────────────────────────────── dashboard ────────────────────

@login_required(login_url="login")
def homepage_view(request):
    total_clients = ClientPersonalInfo.objects.count()

    active_status = (
        ClientStatus.objects
        .exclude(plan="No Plan")
        .filter(status=True, is_cancelled=False)
        .count()
    )

    # FIX: inactive = clients whose only ClientStatus row is still "No Plan"
    inactive_count = ClientStatus.objects.filter(plan="No Plan").count()

    # FIX: completed = fully paid plans (status=False, not cancelled)
    completed_count = (
        ClientStatus.objects
        .exclude(plan="No Plan")
        .filter(status=False, is_cancelled=False)
        .count()
    )

    collected = (
        ClientStatus.objects
        .aggregate(total=Sum("paid_balance"))["total"] or 0
    )

    pending_payments = (
        Payment.objects
        .filter(is_paid=False)
        .exclude(client_status__plan="No Plan")
        .count()
    )

    recent_payments = (
        Payment.objects
        .filter(is_paid=True)
        .select_related("client_status__client", "processed_by")
        .order_by("-date_paid")[:5]
    )

    recent_activity = (
        ActivityLog.objects
        .order_by("-timestamp")[:8]
    )
 
    return render(request, "app/homepage.html", {
        "total_clients":    total_clients,
        "active_status":    active_status,
        "inactive_count":   inactive_count,
        "completed_count":  completed_count,
        "collected":        collected,
        "pending_payments": pending_payments,
        "recent_payments":  recent_payments,
        "recent_activity":  recent_activity,   # ← renamed from recent_logs
    })


# ─────────────────────────────────────────────── clients ──────────────────────

@login_required(login_url="login")
def add_client_view(request):
    if request.method == "POST":
        if not _check_write_role(request, "General Staff"):
            messages.error(request, "Access restricted. Only General Staff can add clients.")
            return redirect("records")
        form                = ClientForm(request.POST)
        beneficiary_formset = BeneficiaryFormSet(request.POST, prefix="beneficiaries")

        if form.is_valid() and beneficiary_formset.is_valid():
            client = form.save(commit=False)

            if ClientPersonalInfo.objects.filter(
                client_first_name__iexact=client.client_first_name,
                client_last_name__iexact=client.client_last_name,
                client_date_birth=client.client_date_birth,
            ).exists():
                form.add_error(None, "A client with this name and date of birth already exists.")
                return render(request, "app/addclient.html", {
                    "form": form, "beneficiary_formset": beneficiary_formset,
                })

            with transaction.atomic():
                client.save()
                beneficiary_formset.instance = client
                beneficiary_formset.save()

                today = datetime.date.today()
                cs = ClientStatus.objects.create(
                    client=client, plan="No Plan",
                    monthly_payment=0, duration=0,
                    months_remaining=0, start_date=today,
                    balance=0, paid_balance=0.00, status=False,
                )
                _generate_payment_rows(cs)

            _log_activity(request.user, "Add Client", f"Client: {client.full_name}")
            messages.success(request, f"Client '{client.full_name}' added successfully.")

            return redirect("records")
    else:
        form                = ClientForm()
        beneficiary_formset = BeneficiaryFormSet(prefix="beneficiaries")

    return render(request, "app/addclient.html", {
        "form": form, "beneficiary_formset": beneficiary_formset,
    })


@login_required(login_url="login")
def records_view(request):
    query   = request.GET.get("q", "").strip()
    clients = ClientPersonalInfo.objects.all().order_by("client_last_name", "client_first_name")

    if query:
        clients = clients.filter(
            Q(client_first_name__icontains=query)     |
            Q(client_middle_name__icontains=query)    |
            Q(client_last_name__icontains=query)      |
            Q(client_contact_number__icontains=query) |
            Q(client_civil_status__icontains=query)   |
            Q(clientstatus__plan__icontains=query)
        ).distinct()

    return render(request, "app/records.html", {"client": clients, "query": query})


@login_required(login_url="login")
def client_details_view(request, pk):
    client      = get_object_or_404(ClientPersonalInfo, pk=pk)
    all_statuses = (
        ClientStatus.objects
        .filter(client=client)
        .exclude(plan="No Plan")
        .order_by("-pk")
    )
    return render(request, "app/client-details.html", {
        "client":       client,
        "all_statuses": all_statuses,
    })


@login_required(login_url="login")
def edit_details_view(request, pk):
    client = get_object_or_404(ClientPersonalInfo, pk=pk)

    if request.method == "POST":
        if not _check_write_role(request, "General Staff"):
            messages.error(request, "Access restricted. Only General Staff can edit client records.")
            return redirect("client-details", pk=pk)
        form                = ClientForm(request.POST, instance=client)
        beneficiary_formset = BeneficiaryFormSet(
            request.POST, instance=client, prefix="beneficiaries"
        )
        if form.is_valid() and beneficiary_formset.is_valid():
            with transaction.atomic():
                form.save()
                beneficiary_formset.save()
            _log_activity(request.user, "Edit Record", f"Client: {client.full_name}")
            messages.success(request, "Client updated successfully.")
            return redirect("records")
    else:
        form                = ClientForm(instance=client)
        beneficiary_formset = BeneficiaryFormSet(instance=client, prefix="beneficiaries")

    return render(request, "app/edit_details.html", {
        "form": form, "beneficiary_formset": beneficiary_formset, "client": client,
    })


@login_required(login_url="login")
def delete_client_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Method not allowed."}, status=405)
    
    if not _check_write_role(request, "General Staff"):           # ← ADD THIS
        return JsonResponse({"success": False, "error": "Access restricted."})

    if not _check_pin(request.user, request.POST.get("pin", "")):
        return JsonResponse({"success": False, "error": "Invalid PIN."})

    client = get_object_or_404(ClientPersonalInfo, pk=pk)
    name   = client.full_name
    client.delete()
    _log_activity(request.user, "Delete Client", f"Client: {name}")
    return JsonResponse({"success": True, "name": name})


# ─────────────────────────────────────────────── payments ─────────────────────

@_require_role("General Staff", "Financial Staff")
def add_payment_view(request, pk):
    client      = get_object_or_404(ClientPersonalInfo, pk=pk)
    all_statuses = (
        ClientStatus.objects
        .filter(client=client)
        .exclude(plan="No Plan")
        .order_by("-pk")
    )
    for cs in all_statuses:
        _generate_payment_rows(cs)
    return render(request, "app/add_payment.html", {
        "client":       client,
        "all_statuses": all_statuses,
    })


@_require_role("General Staff", "Financial Staff")
def payment_history_view(request, pk):
    client      = get_object_or_404(ClientPersonalInfo, pk=pk)
    all_statuses = (
        ClientStatus.objects
        .filter(client=client)
        .exclude(plan="No Plan")
        .order_by("-pk")
    )
 
    if not all_statuses.exists():
        messages.info(request, "This client has no plan assigned yet.")
        return redirect("add-payment", pk=pk)
 
    # ── POST: process a payment ───────────────────────────────────────────────
    if request.method == "POST":
        if not _check_write_role(request, "Financial Staff"):
            return JsonResponse({"success": False, "error": "Access restricted. Only Financial Staff can record payments."})
        payment_id = request.POST.get("payment_id", "")
        pin        = request.POST.get("pin", "")
 
        if not _check_pin(request.user, pin):
            return JsonResponse({"success": False, "error": "Invalid PIN."})
 
        # Find the payment tied to ANY of this client's plans
        payment       = get_object_or_404(Payment, pk=payment_id, client_status__client=client)
        client_status = payment.client_status
 
        if payment.is_paid:
            return JsonResponse({"success": False, "error": "This month is already paid."})
 
        with transaction.atomic():
            payment.is_paid   = True
            payment.date_paid = timezone.now()
            payment.processed_by = request.user
            payment.save()
 
            client_status.paid_balance    += payment.amount
            client_status.balance         -= payment.amount
            client_status.months_remaining = client_status.payments.filter(is_paid=False).count()
            client_status.date_paid        = payment.date_paid
 
            if client_status.months_remaining == 0:
                client_status.status       = False
                client_status.date_fully_paid = datetime.date.today()
 
            client_status.save()
 
        _log_activity(
            request.user, "Add Payment",
            f"Client: {client.full_name} — {payment.month.strftime('%B %Y')} — ₱{payment.amount}"
        )
 
        return JsonResponse({
            "success":          True,
            "date_paid":        localtime(payment.date_paid).strftime("%b %d, %Y %I:%M %p"),
            "paid_balance":     str(client_status.paid_balance),
            "balance":          str(client_status.balance),
            "months_remaining": client_status.months_remaining,
        })
 
    # ── GET: select which plan to display ────────────────────────────────────
    selected_pk = request.GET.get("plan")
    if selected_pk:
        try:
            client_status = all_statuses.get(pk=selected_pk)
        except ClientStatus.DoesNotExist:
            client_status = _get_active_status(client) or all_statuses.first()
    else:
        client_status = _get_active_status(client) or all_statuses.first()
 
    _generate_payment_rows(client_status)
    payments = client_status.payments.all()
 
    return render(request, "app/payment_history.html", {
        "client":        client,
        "client_status": client_status,
        "all_statuses":  all_statuses,
        "payments":      payments,
    })


# ─────────────────────────────────────────────── plan ─────────────────────────

@login_required(login_url="login")
def plan(request, pk):
    client = get_object_or_404(ClientPersonalInfo, pk=pk)
    plans  = ClientStatus.objects.filter(client=client)
    form   = PlanForm()

    if request.method == "POST":
        if not _check_write_role(request, "Financial Staff"):
            messages.error(request, "Access restricted. Only Financial Staff can assign plans.")
            return redirect("plan", pk=pk)
        form = PlanForm(request.POST)
        if form.is_valid():
            d         = form.cleaned_data
            plan_name = d["plan"]
            monthly   = d["monthly_payment"]
            duration  = d["duration"]
            total     = monthly * duration
            today     = datetime.date.today()

            down_payment     = d.get("down_payment")
            discount_percent = d.get("discount_percent") or 0
            phase            = d.get("phase", "").strip() or None
            block            = d.get("block", "").strip() or None
            section          = d.get("section", "").strip() or None
            lot_number       = d.get("lot_number", "").strip() or None
            pa_number        = d.get("pa_number", "").strip() or None
            # FIX: new tracking fields
            contract_number  = d.get("contract_number", "").strip() or None
            interment_date   = d.get("interment_date")
            pa_date          = d.get("pa_date")
            # Columbarium / THS / THTC fields
            column_level      = d.get("column_level", "").strip() or None
            columbarium_type  = d.get("columbarium_type", "").strip() or None
            columbarium_level = d.get("columbarium_level") or None
            tomb_number       = d.get("tomb_number", "").strip() or None

            # FIX: only block if there is an ACTIVE, non-cancelled plan.
            # Cancelled or completed plans do not block a new assignment.
            blocking = plans.exclude(plan="No Plan").filter(
                is_cancelled=False, status=True
            )
            if blocking.exists():
                messages.error(
                    request,
                    "This client already has an active plan. "
                    "Cancel or complete it before assigning a new one."
                )
                return render(request, "app/plan.html", {
                    "client": client, "plans": plans, "form": form,
                })

            no_plan = plans.filter(plan="No Plan").first()

            def _apply(cs):
                if cs.pk:
                    cs.payments.all().delete()
                cs.plan             = plan_name
                cs.monthly_payment  = monthly
                cs.duration         = duration
                cs.months_remaining = duration
                cs.start_date       = today
                cs.balance          = total
                cs.paid_balance     = 0
                cs.status           = True
                cs.date_paid        = None
                cs.down_payment     = down_payment
                cs.discount_percent = discount_percent
                cs.phase            = phase
                cs.block            = block
                cs.section          = section
                cs.lot_number       = lot_number
                cs.pa_number        = pa_number
                cs.contract_number  = contract_number
                cs.interment_date   = interment_date
                cs.pa_date          = pa_date
                cs.column_level      = column_level
                cs.columbarium_type  = columbarium_type
                cs.columbarium_level = columbarium_level
                cs.tomb_number       = tomb_number
                cs.is_cancelled     = False
                cs.cancellation_reason = None
                cs.cancellation_date   = None
                cs.date_fully_paid     = None
                cs.save()
                _generate_payment_rows(cs)

            if no_plan:
                _apply(no_plan)
            else:
                new_cs = ClientStatus(client=client)
                _apply(new_cs)

            _log_activity(request.user, "Add Plan", f"Client: {client.full_name} — Plan: {plan_name}")
            messages.success(request, f"Plan '{plan_name}' has been assigned successfully.")
            return redirect("plan", pk=pk)

    return render(request, "app/plan.html", {
        "client": client, "plans": plans, "form": form,
    })


# ─────────────────────────────────────────────── cancel plan (NEW) ────────────

@login_required(login_url="login")
def cancel_plan_view(request, pk):
    """Cancel an active lot plan. pk = ClientStatus.pk"""
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Method not allowed."}, status=405)
    
    if not _check_write_role(request, "Financial Staff"):          # ← ADD THIS
        return JsonResponse({"success": False, "error": "Access restricted."})

    if not _check_pin(request.user, request.POST.get("pin", "")):
        return JsonResponse({"success": False, "error": "Invalid PIN."})

    cs = get_object_or_404(ClientStatus, pk=pk)

    if cs.is_cancelled:
        return JsonResponse({"success": False, "error": "Plan is already cancelled."})
    if cs.plan == "No Plan":
        return JsonResponse({"success": False, "error": "No active plan to cancel."})

    reason = request.POST.get("reason", "").strip() or "No reason provided."
    cs.status              = False
    cs.is_cancelled        = True
    cs.cancellation_reason = reason
    cs.cancellation_date   = datetime.date.today()
    cs.save()

    _log_activity(request.user, "Cancel Plan", f"Plan ID: {pk} — {cs.plan}")
    return JsonResponse({"success": True})


# ─────────────────────────────────────────────── lots ─────────────────────────

@login_required(login_url="login")
def lots_view(request):
    lots = (
        ClientStatus.objects
        .exclude(plan="No Plan")
        .select_related("client")
        .order_by("phase", "block", "section", "lot_number")
    )

    q_type              = request.GET.get("lot_type", "").strip()
    q_phase             = request.GET.get("phase", "").strip()
    q_block             = request.GET.get("block", "").strip()
    q_section           = request.GET.get("section", "").strip()
    q_lot               = request.GET.get("lot_number", "").strip()
    q_status            = request.GET.get("status", "").strip()
    q_column_level      = request.GET.get("column_level", "").strip()
    q_columbarium_type  = request.GET.get("columbarium_type", "").strip()
    q_columbarium_level = request.GET.get("columbarium_level", "").strip()
    q_tomb_number       = request.GET.get("tomb_number", "").strip()

    if q_type:
        lots = lots.filter(plan__icontains=q_type)
    if q_phase:
        lots = lots.filter(phase__icontains=q_phase)
    if q_block:
        lots = lots.filter(block__icontains=q_block)
    if q_section:
        lots = lots.filter(section__icontains=q_section)
    if q_lot:
        lots = lots.filter(lot_number__icontains=q_lot)
    if q_column_level:
        lots = lots.filter(column_level__icontains=q_column_level)
    if q_columbarium_type:
        lots = lots.filter(columbarium_type=q_columbarium_type)
    if q_columbarium_level:
        try:
            lots = lots.filter(columbarium_level=int(q_columbarium_level))
        except ValueError:
            pass
    if q_tomb_number:
        lots = lots.filter(tomb_number__icontains=q_tomb_number)
    if q_status == "Active":
        lots = lots.filter(status=True, is_cancelled=False)
    elif q_status == "Completed":
        lots = lots.filter(status=False, is_cancelled=False)
    elif q_status == "Cancelled":
        lots = lots.filter(is_cancelled=True)

    plan_choices = [c[0] for c in ClientStatus.PLAN_CHOICES if c[0] != "No Plan"]

    return render(request, "app/lots.html", {
        "lots":                 lots,
        "q_type":               q_type,
        "q_phase":              q_phase,
        "q_block":              q_block,
        "q_section":            q_section,
        "q_lot":                q_lot,
        "q_status":             q_status,
        "q_column_level":       q_column_level,
        "q_columbarium_type":   q_columbarium_type,
        "q_columbarium_level":  q_columbarium_level,
        "q_tomb_number":        q_tomb_number,
        "plan_choices":         plan_choices,
    })

# ─────────────────────────────────────────────── bookings ─────────────────────

@login_required(login_url="login")
def bookings_view(request):
    # Auto-complete bookings whose date has passed
    from django.utils import timezone
    today = timezone.localdate()
    Booking.objects.filter(status="Active", booking_date__lt=today).update(status="Completed")
 
    form     = BookingForm()
    bookings = Booking.objects.order_by("-booking_date", "booking_time")
 
    if request.method == "POST":
        action = request.POST.get("action", "")
 
        # ── Mark as No Show ───────────────────────────────────────────────────
        if action == "no_show":
            if not _check_write_role(request, "General Staff"):
                return JsonResponse({"success": False, "error": "Access restricted."})
            if not _check_pin(request.user, request.POST.get("pin", "")):
                return JsonResponse({"success": False, "error": "Invalid PIN."})
            booking = get_object_or_404(Booking, pk=request.POST.get("pk"))
            booking.status              = "No Show"
            booking.cancellation_reason = request.POST.get("reason", "").strip() or "Client did not show up."
            booking.cancelled_at        = timezone.now()
            booking.save()
            _log_activity(request.user, "Cancel Booking",
                f"{booking.client_name} — No Show on {booking.booking_date}")
            return JsonResponse({"success": True})
 
        # ── Add new booking ───────────────────────────────────────────────────
        if not _check_write_role(request, "General Staff"):
            messages.error(request, "Access restricted. Only General Staff can add bookings.")
            return redirect("bookings")
        form = BookingForm(request.POST)
        if form.is_valid():
            booking = form.save()
            _log_activity(
                request.user, "Add Booking",
                f"{booking.client_name} on {booking.booking_date} "
                f"at {booking.get_booking_time_display()}"
            )
            messages.success(
                request,
                f"Booking saved for {booking.client_name} on "
                f"{booking.booking_date} at {booking.get_booking_time_display()}."
            )
            return redirect("bookings")
 
    # Build booked slots dict for calendar JS
    active_bookings = (
        Booking.objects
        .filter(status="Active")
        .values("booking_date", "booking_time")
    )
    booked_slots: dict = defaultdict(list)
    for b in active_bookings:
        date_str = b["booking_date"].strftime("%Y-%m-%d")
        booked_slots[date_str].append(b["booking_time"])
 
    return render(request, "app/bookings.html", {
        "form":              form,
        "bookings":          bookings,
        "booked_slots_json": json.dumps(dict(booked_slots)),
    })


# ─────────────────────────────────────────────── cancel booking (NEW) ─────────

@login_required(login_url="login")
def cancel_booking_view(request, pk):
    """Cancel an active booking. pk = Booking.pk"""
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Method not allowed."}, status=405)
    
    if not _check_write_role(request, "General Staff"):            # ← ADD THIS
        return JsonResponse({"success": False, "error": "Access restricted."})

    if not _check_pin(request.user, request.POST.get("pin", "")):
        return JsonResponse({"success": False, "error": "Invalid PIN."})

    booking = get_object_or_404(Booking, pk=pk)

    if booking.status == "Cancelled":
        return JsonResponse({"success": False, "error": "Booking is already cancelled."})

    reason = request.POST.get("reason", "").strip() or "No reason provided."
    booking.status              = "Cancelled"
    booking.cancellation_reason = reason
    booking.cancelled_at        = timezone.now()
    booking.save()

    _log_activity(request.user, "Cancel Booking",
    f"Booking ID: {pk} — {booking.client_name}")
    return JsonResponse({"success": True})


# ─────────────────────────────────────────────── calendar ─────────────────────

@login_required(login_url="login")
def calendar_view(request):
    from collections import Counter

    today = datetime.date.today()

    try:
        year  = int(request.GET.get("year",  today.year))
        month = int(request.GET.get("month", today.month))
    except ValueError:
        year, month = today.year, today.month

    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1

    month_bookings = Booking.objects.filter(
        booking_date__year=year,
        booking_date__month=month,
        status="Active",
    )

    total_slots    = len(Booking.TIME_SLOTS)  # 11
    booking_counts = Counter(b.booking_date.day for b in month_bookings)

    fully_booked_dates    = {d for d, c in booking_counts.items() if c >= total_slots}
    partially_booked_dates = {d for d in booking_counts if d not in fully_booked_dates}

    cal_weeks = cal_module.monthcalendar(year, month)

    prev_month, prev_year = (12, year - 1) if month == 1 else (month - 1, year)
    next_month, next_year = (1,  year + 1) if month == 12 else (month + 1, year)

    selected_date_str = request.GET.get("date", "").strip()
    selected_date     = None
    booked_times      = set()

    if selected_date_str:
        try:
            selected_date = datetime.date.fromisoformat(selected_date_str)
            booked_times  = set(
                Booking.objects.filter(
                    booking_date=selected_date,
                    status="Active",
                ).values_list("booking_time", flat=True)
            )
        except ValueError:
            selected_date_str = ""

    return render(request, "app/calendar.html", {
        "cal_weeks":              cal_weeks,
        "month":                  month,
        "year":                   year,
        "month_name":             cal_module.month_name[month],
        "booked_dates":           fully_booked_dates,
        "partially_booked_dates": partially_booked_dates,
        "prev_month":             prev_month,
        "prev_year":              prev_year,
        "next_month":             next_month,
        "next_year":              next_year,
        "today":                  today,
        "time_slots":             Booking.TIME_SLOTS,
        "booked_times":           booked_times,
        "selected_date":          selected_date,
        "selected_date_str":      selected_date_str,
        "day_names":              ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
    })


# ─────────────────────────────────────────────── monitor ──────────────────────

@_admin_required
def monitor_view(request):
    logs = ActivityLog.objects.select_related("user").order_by("-timestamp")
 
    q_name   = request.GET.get("name",      "").strip()
    q_action = request.GET.get("action",    "").strip()
    q_from   = request.GET.get("date_from", "").strip()
    q_to     = request.GET.get("date_to",   "").strip()
 
    if q_name:
        logs = logs.filter(staff_name__icontains=q_name)
    if q_action and q_action != "All":
        logs = logs.filter(action=q_action)
    if q_from:
        logs = logs.filter(timestamp__date__gte=q_from)
    if q_to:
        logs = logs.filter(timestamp__date__lte=q_to)
 
    return render(request, "app/monitor.html", {
        "logs":     logs,
        "q_name":   q_name,
        "q_action": q_action,
        "q_from":   q_from,
        "q_to":     q_to,
    })


# ─────────────────────────────────────────────── employees ────────────────────

@_admin_required
def employee_view(request):
    query     = request.GET.get("q", "").strip()
    employees = UserLog.objects.select_related("user").all().order_by("last_name", "first_name")

    if query:
        employees = employees.filter(
            Q(first_name__icontains=query)  |
            Q(last_name__icontains=query)   |
            Q(middle_name__icontains=query) |
            Q(role__icontains=query)        |
            Q(phone_number__icontains=query)
        ).distinct()

    return render(request, "app/viewemployee.html", {
        "employees": employees, "query": query,
    })


@_admin_required
def add_employee_view(request):
    if request.method == "POST":
        form = EmployeeCreateForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data

            if d["username"].lower() == "admin":
                form.add_error("username", "The username 'admin' is reserved.")
                return render(request, "app/add-employee.html", {"form": form})

            if User.objects.filter(username__iexact=d["username"]).exists():
                form.add_error("username", "Username already exists.")
                return render(request, "app/add-employee.html", {"form": form})

            with transaction.atomic():
                user = User.objects.create_user(
                    username=d["username"],
                    password=d["password"],
                    email=d.get("email", ""),
                    first_name=d["first_name"],
                    last_name=d["last_name"],
                )
                UserLog.objects.create(
                    user=user,
                    role=d["role"],
                    first_name=d["first_name"],
                    middle_name=d.get("middle_name", "") or "",
                    last_name=d["last_name"],
                    date_of_birth=d["date_of_birth"],
                    government_id=d.get("government_id", ""),
                    phone_number=d.get("phone_number", ""),
                    email=d.get("email", ""),
                    address=d["address"],
                    emergency_contact_name=d["emergency_contact_name"],
                    emergency_contact_number=d["emergency_contact_number"],
                    activities="Account created",
                    pin=make_password(d["pin"]),
                )

            messages.success(
                request,
                f"Employee '{d['first_name']} {d['last_name']}' created successfully."
            )
            _log_activity(
                request.user, "Add Employee",
                f"Employee: {d['first_name']} {d['last_name']} ({d['role']})"
            )
            return redirect("employee")

        return render(request, "app/add-employee.html", {"form": form})

    return render(request, "app/add-employee.html", {"form": EmployeeCreateForm()})


@_admin_required
def details_employee_view(request, pk):
    employee = get_object_or_404(UserLog, pk=pk)

    duration_str = "—"
    if employee.time_in and employee.time_out:
        delta   = employee.time_out - employee.time_in
        minutes = int(delta.total_seconds() // 60)
        duration_str = f"{minutes // 60}h {minutes % 60}m"

    return render(request, "app/employee-details.html", {
        "employee": employee, "duration_str": duration_str,
    })


@_admin_required
def edit_employee_view(request, pk):
    employee = get_object_or_404(UserLog, pk=pk)

    if request.method == "POST":
        form = EmployeeUpdateForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data
            with transaction.atomic():
                employee.first_name               = d["first_name"]
                employee.middle_name              = d.get("middle_name", "") or ""
                employee.last_name                = d["last_name"]
                employee.date_of_birth            = d["date_of_birth"]
                employee.address                  = d["address"]
                employee.email                    = d["email"]
                employee.phone_number             = d["phone_number"]
                employee.emergency_contact_name   = d["emergency_contact_name"]
                employee.emergency_contact_number = d["emergency_contact_number"]
                employee.role                     = d["role"]
                employee.government_id            = d["government_id"]
                employee.pin                      = make_password(d["pin"])
                employee.save()

                if employee.user:
                    employee.user.first_name = d["first_name"]
                    employee.user.last_name  = d["last_name"]
                    employee.user.email      = d["email"]
                    if d.get("new_password"):
                        employee.user.set_password(d["new_password"])
                    employee.user.save()

            messages.success(request, "Employee updated successfully.")
            _log_activity(
                request.user, "Edit Employee",
                f"Employee: {employee.full_name}"
            )
            return redirect("details-employee", pk=pk)
    else:
        form = EmployeeUpdateForm(initial={
            "first_name":               employee.first_name,
            "middle_name":              employee.middle_name,
            "last_name":                employee.last_name,
            "date_of_birth":            employee.date_of_birth,
            "address":                  employee.address,
            "email":                    employee.email,
            "phone_number":             employee.phone_number,
            "emergency_contact_name":   employee.emergency_contact_name,
            "emergency_contact_number": employee.emergency_contact_number,
            "role":                     employee.role,
            "government_id":            employee.government_id,
        })

    return render(request, "app/employee-edit.html", {
        "form": form, "employee": employee,
    })


@_admin_required
def delete_employee_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Method not allowed."}, status=405)

    if not _check_pin(request.user, request.POST.get("pin", "")):
        return JsonResponse({"success": False, "error": "Invalid PIN."})

    employee = get_object_or_404(UserLog, pk=pk)
    if employee.user == request.user:
        return JsonResponse({"success": False, "error": "You cannot delete your own account."})

    name = employee.full_name
    with transaction.atomic():
        if employee.user:
            employee.user.delete()
        else:
            employee.delete()

    _log_activity(request.user, "Delete Employee", f"Employee: {name}")
    return JsonResponse({"success": True, "name": name})


def _xl_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
 
 
def _xl_header_row(ws, headers):
    """Append a styled dark header row and return the row number."""
    ws.append(headers)
    row = ws.max_row
    dark  = PatternFill(start_color="2F3640", end_color="2F3640", fill_type="solid")
    white = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = dark
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20
    return row
 
 
def _xl_auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter  = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 4, 55)


_admin_required
@require_POST
def backup_now_view(request):
    """
    Manual backup trigger — admin only.
    Called by the Backup Now button on the Monitor page.
    Runs synchronously so the result is returned immediately.
    """
    from .backup import trigger_manual_backup
    result = trigger_manual_backup()
    return JsonResponse(result)
 
 
@_admin_required
def backup_status_view(request):
    """
    Returns the list of local backup files for display on Monitor page.
    """
    from django.conf import settings as django_settings
    from pathlib import Path
 
    backup_dir = Path(
        getattr(django_settings, "BACKUP_OFFLINE_DIR",
                str(django_settings.BASE_DIR / "backups"))
    )
 
    files = []
    if backup_dir.exists():
        all_zips = sorted(backup_dir.glob("backup_*.zip"), reverse=True)
        for f in all_zips[:15]:          # show last 15 only
            stat = f.stat()
            files.append({
                "name":     f.name,
                "size_kb":  round(stat.st_size / 1024, 1),
                "modified": datetime.datetime.fromtimestamp(stat.st_mtime)
                            .strftime("%b %d, %Y %I:%M %p"),
            })
        last = all_zips[0] if all_zips else None
        last_info = {
            "filename": last.name,
            "modified": datetime.datetime.fromtimestamp(last.stat().st_mtime)
                        .strftime("%b %d, %Y %I:%M %p"),
        } if last else None
    else:
        last_info = None
 
    return JsonResponse({
        "files":     files,
        "directory": str(backup_dir),
        "last":      last_info,
        "total":     len(files),
    })


@_admin_required
def system_settings_view(request):
    """
    Admin-only page to manage encrypted system secrets.
    Values are stored encrypted in DB — never exposed in source code.
    """
    from .models import SystemSecret
 
    # ── Default secrets with labels (created if they don't exist) ─────────────
    DEFAULTS = [
        ("BACKUP_EMAIL_HOST",     "smtp.gmail.com", "Email SMTP Host"),
        ("BACKUP_EMAIL_PORT",     "587",            "Email SMTP Port"),
        ("BACKUP_EMAIL_USER",     "",               "Email Sender Address"),
        ("BACKUP_EMAIL_PASSWORD", "",               "Email App Password"),
        ("BACKUP_EMAIL_TO",       "",               "Email Recipient Address"),
        ("ADMIN_PIN",             "",               "Admin PIN (4 digits)"),
    ]
 
    for key, default_val, label in DEFAULTS:
        SystemSecret.objects.get_or_create(
            key=key,
            defaults={"label": label, "encrypted_value": ""},
        )
 
    message = None
    error   = None
 
    if request.method == "POST":
        pin = request.POST.get("pin", "")
        if pin != settings.ADMIN_PIN:
            error = "Incorrect PIN."
        else:
            for key, _, _ in DEFAULTS:
                new_val = request.POST.get(key, "").strip()
                if new_val:   # only update if a value was provided
                    obj = SystemSecret.objects.get(key=key)
                    obj.set_value(new_val)
                    obj.save()
            _log_activity(request.user, "System Settings", "Updated encrypted system secrets.")
            message = "Settings saved and encrypted successfully."
 
    # Build context — decrypt current values for display (masked)
    secrets = []
    for key, _, label in DEFAULTS:
        obj = SystemSecret.objects.get(key=key)
        val = obj.get_value()
        secrets.append({
            "key":   key,
            "label": label,
            "set":   bool(val),   # just show whether it's set, not the value
        })
 
    return render(request, "app/system_settings.html", {
        "secrets": secrets,
        "message": message,
        "error":   error,
    })
 
 
# ── Bookings export ───────────────────────────────────────────────────────────
 
@login_required(login_url="login")
def export_bookings_excel(request):
    from django.utils import timezone
    today = timezone.localdate()
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"
 
    # Title
    ws.append(["GREEN GARDEN — BOOKING RECORDS"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.append([f"Exported: {today.strftime('%B %d, %Y')}"])
    ws.append([])
 
    _xl_header_row(ws, [
        "#", "Client Name", "Contact Number", "Event Type",
        "Booking Date", "Time", "Notes", "Status",
    ])
 
    bookings = Booking.objects.order_by("-booking_date", "booking_time")
    for i, b in enumerate(bookings, 1):
        ws.append([
            i,
            b.client_name,
            b.contact_number,
            b.event_type,
            b.booking_date.strftime("%B %d, %Y"),
            b.get_booking_time_display(),
            b.notes or "",
            b.status,
        ])
 
    ws.freeze_panes = "A5"
    _xl_auto_width(ws)
    return _xl_response(wb, f"bookings_{today}.xlsx")
 
 
# ── Payment History export ────────────────────────────────────────────────────
 
@login_required(login_url="login")
def export_payment_history_excel(request, pk):
    import datetime
    client      = get_object_or_404(ClientPersonalInfo, pk=pk)
    all_statuses = (
        ClientStatus.objects
        .filter(client=client)
        .exclude(plan="No Plan")
        .order_by("-pk")
    )
 
    selected_pk = request.GET.get("plan")
    if selected_pk:
        try:
            client_status = all_statuses.get(pk=selected_pk)
        except ClientStatus.DoesNotExist:
            client_status = _get_active_status(client) or all_statuses.first()
    else:
        client_status = _get_active_status(client) or all_statuses.first()
 
    if not client_status:
        messages.info(request, "No plan found for this client.")
        return redirect("payment-history", pk=pk)
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment History"
 
    today = datetime.date.today().strftime("%B %d, %Y")
 
    # Title
    ws.append(["GREEN GARDEN — PAYMENT HISTORY"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.append([f"Exported: {today}"])
    ws.append([])
 
    # Client info section
    ws.append(["CLIENT INFORMATION"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
    info_rows = [
        ("ID",            client.pk),
        ("Full Name",     client.full_name),
        ("Contact",       client.client_contact_number),
        ("Civil Status",  client.client_civil_status),
        ("Date of Birth", str(client.client_date_birth)),
        ("Address",       client.client_address),
    ]
    for label, value in info_rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
 
    ws.append([])
 
    # Plan info section
    ws.append(["PLAN INFORMATION"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
    plan_rows = [
        ("Plan",           client_status.plan),
        ("Monthly",        f"₱{client_status.monthly_payment:,.2f}"),
        ("Duration",       f"{client_status.duration} months"),
        ("Start Date",     str(client_status.start_date)),
        ("Down Payment",   f"₱{client_status.down_payment:,.2f}" if client_status.down_payment else "—"),
        ("Discount",       f"{client_status.discount_percent}%" if client_status.discount_percent else "0%"),
        ("Total Paid",     f"₱{client_status.paid_balance:,.2f}"),
        ("Balance",        f"₱{client_status.balance:,.2f}"),
        ("Date Fully Paid", str(client_status.date_fully_paid) if client_status.date_fully_paid else "—"),
        ("Status",         "Cancelled" if client_status.is_cancelled else ("Active" if client_status.status else "Completed")),
    ]
    for label, value in plan_rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
 
    ws.append([])
 
    # Payment schedule
    ws.append(["MONTHLY PAYMENT SCHEDULE"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
 
    _xl_header_row(ws, ["#", "Month", "Amount", "Status", "Date Paid", "Processed By"])
 
    for i, p in enumerate(client_status.payments.all(), 1):
        processed_by = "—"
        if p.processed_by:
            if p.processed_by.is_superuser or p.processed_by.username == "admin":
                processed_by = "Admin"
            else:
                try:
                    processed_by = p.processed_by.userlog.full_name or p.processed_by.username
                except Exception:
                    processed_by = p.processed_by.username
 
        ws.append([
            i,
            p.month.strftime("%B %Y"),
            float(p.amount),
            "Paid" if p.is_paid else "Unpaid",
            p.date_paid.strftime("%B %d, %Y %I:%M %p") if p.date_paid else "—",
            processed_by,
        ])
 
    ws.freeze_panes = f"A{ws.max_row - len(list(client_status.payments.all()))}"
    _xl_auto_width(ws)
    return _xl_response(wb, f"payment_history_{client.pk}_{client_status.plan}.xlsx")
 
 
# ── Lots export ───────────────────────────────────────────────────────────────
 
@login_required(login_url="login")
def export_lots_excel(request):
    import datetime
    today = datetime.date.today().strftime("%B %d, %Y")
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lots"
 
    ws.append(["GREEN GARDEN — LOT RECORDS"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.append([f"Exported: {today}"])
    ws.append([])
 
    _xl_header_row(ws, [
        "PHASE", "BLOCK", "SECTION", "LOT NUMBER", "LOT TYPE",
        "NAME OF BUYER", "DATE OF SALES", "DATE REGISTERED",
        "C.O NUMBER", "PA NUMBER", "MONTHLY", "BALANCE", "STATUS",
    ])
 
    lots = (
        ClientStatus.objects
        .exclude(plan="No Plan")
        .select_related("client")
        .order_by("phase", "block", "section", "lot_number")
    )
 
    for lot in lots:
        status = (
            "Cancelled" if lot.is_cancelled
            else ("Active" if lot.status else "Completed")
        )
        ws.append([
            lot.phase or "—",
            lot.block or "—",
            lot.section or "—",
            lot.lot_number or "—",
            lot.plan,
            lot.client.full_name,
            lot.start_date.strftime("%B %d, %Y") if lot.start_date else "—",
            lot.pa_date.strftime("%B %d, %Y") if lot.pa_date else "—",
            lot.contract_number or "—",
            lot.pa_number or "—",
            float(lot.monthly_payment),
            float(lot.balance),
            status,
        ])
 
    ws.freeze_panes = "A5"
    _xl_auto_width(ws)
    return _xl_response(wb, f"lots_{datetime.date.today()}.xlsx")
 
 
# ── Client Records export ─────────────────────────────────────────────────────
 
@login_required(login_url="login")
def export_records_excel(request):
    import datetime
    today = datetime.date.today().strftime("%B %d, %Y")
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Records"
 
    ws.append(["GREEN GARDEN — CLIENT RECORDS"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.append([f"Exported: {today}"])
    ws.append([])
 
    _xl_header_row(ws, [
        "ID", "FIRST NAME", "MIDDLE NAME", "LAST NAME",
        "ADDRESS", "CONTACT", "CIVIL STATUS", "DATE OF BIRTH",
        "RELIGION", "OCCUPATION", "EMPLOYER", "PLAN", "STATUS",
    ])
 
    clients = ClientPersonalInfo.objects.all().order_by("client_last_name")
    for c in clients:
        latest = ClientStatus.objects.filter(client=c).order_by("-pk").first()
        plan   = latest.plan if latest else "No Plan"
        status = (
            "Cancelled" if (latest and latest.is_cancelled)
            else ("Active" if (latest and latest.status) else "Completed")
        )
        ws.append([
            c.pk,
            c.client_first_name,
            c.client_middle_name or "",
            c.client_last_name,
            c.client_address,
            c.client_contact_number,
            c.client_civil_status,
            str(c.client_date_birth),
            c.client_religion,
            c.client_occupation,
            c.client_employer_name,
            plan,
            status,
        ])
 
    ws.freeze_panes = "A5"
    _xl_auto_width(ws)
    return _xl_response(wb, f"client_records_{datetime.date.today()}.xlsx")
